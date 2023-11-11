"""
This script performs incremental updates about client's CTA settle, recording them into a csv file.
Some part of it may contain non-disclosable information of my previous employer, such as the netdisk's folder names.
Therefore, I replaced some non-disclosable information by "#", thank you for understanding!
"""
#import
import sys
import os
import subprocess
import pandas as pd
import numpy as np
import openpyxl


def update(fund_name, current_path, path_raw_data, path_split_data):
    # part 1
    # I used the openpyxl package
    # The advantage of this package is that it can read as well as write Excel
    # even though I used pandas to read, I still have to use openpyxl to write

    if fund_name == "#####":
        file_path = f"#####.xlsx"
        write_main_table_book = openpyxl.load_workbook(f"#####.xlsx")
    else:
        print("Error, fund not found")
        exit()

    # The file you have just opened is in the form of a workbook, so be precise about the sheet here.
    write_main_table = write_main_table_book.worksheets[0]

    # getting the number of rows is useful at the last setp
    numRow = write_main_table.max_row
    # Get the first line of information for the location of the single account
    # Here's a bit of openpyxl logic. list(table.rows) is a big list containing information about each row, 
    # and each element in this big list is a tuple about each row.
    # Each element in the list is a tuple about each row, 
    # and each element in the tuple is in the form of a cell. 
    # cell.value is the original value.
    # For example, if I want to get the information of row i and column j, 
    # it is list(table.rows)[i][j].value, the same as below.
    header = list(cell.value for cell in list(write_main_table.rows)[0])

    # This is used here for insurance purposes to determine the location of the last useful line.
    for i in range(-1, -numRow - 1, -1):  # From back to front, line by line
        if list(write_main_table.rows)[i][0].value != None:  # When there's something in the "date" of this line, record the position.
            # Instead of recording the position as a negative number from back to front, 
            # you add the total number of rows to make it an index from front to back.
            last_date_pos = numRow + i  # This position is used to get the previous day's access amount
            try:
                last_date_list = list(write_main_table.rows)[i][0].value.split("/")
                last_date_list[1] = f"{last_date_list[1]:0>2}"
                last_date_list[2] = f"{last_date_list[2]:0>2}"
                last_date = "-".join(last_date_list)
            except:
                last_date = str(list(write_main_table.rows)[i][0].value)[:10]
            break

    path_cta = f"{path_raw_data}/{fund_name}"
    users_of_cta = os.listdir(path_cta)
    users_of_cta.sort()
    users_of_cta = users_of_cta[1:]
    # Create a new list that is based on the sort order in Excel
    new_users_of_cta = []

    # First go through the user obtained from the crawler, 
    # if it is not in Excel then it means that it is a new customer and needs to be added.
    for user in users_of_cta:
        if user not in header:
            # write the new user's space
            write_main_table.cell(1, len(header)+1, user)
            write_main_table.cell(1, len(header)+2, "")
            write_main_table.cell(1, len(header)+3, "")
            write_main_table.cell(1, len(header)+4, "")
            # you must save it
            write_main_table_book.save(file_path)
            # update header
            header.extend([user, "", "", ""])
    # Then you need to put the users in the order of the header in order to add the data
    for col in header:
        if col in users_of_cta:
            new_users_of_cta.append(col)

    # Here's the loop to get the date of each fund, and this loop is to get the latest date.
    # The reason it's so tricky is because the files in the folder are unordered, so you have to get them all and sort them.
    # It doesn't matter if it's a crawler or an account_summary, because it doesn't open the file or anything.
    new_all_dayfiles_cta=[]
    for user in users_of_cta:
        try:
            each_path = f"{path_cta}/{user}"
            new_all_dayfiles_cta.extend([datefile.split(".")[0] for datefile in os.listdir(each_path)])
        except:
            each_path = f"{path_split_data}/{user}/account_summary"
            new_all_dayfiles_cta.extend([datefile.split(".")[0] for datefile in os.listdir(each_path)])
        new_all_dayfiles_cta = list(set(new_all_dayfiles_cta))
    new_all_dayfiles_cta.sort()

    # The informations are the data of the new row to be added, first defaulting to all zeros, 
    # and then replacing the numbers according to the corresponding grid
    # The reason for doing this is because there will be some customers in Excel 
    # who don't have open positions that aren't shown in the crawler, and their positions need to be left open.
    # If you feel it's not quite standardized to write it as 0 here, replace it with '
    informations = [0 for i in range(len(header))]

    all_difference = new_all_dayfiles_cta[new_all_dayfiles_cta.index(last_date)+1:]

    money_in = float(list(write_main_table.rows)[last_date_pos][2].value)

    # go over clients
    for difference in all_difference:
        settle = 0
        commission = 0
        margin_occupied = 0
        for users in new_users_of_cta:
            print(users)
            # First determine if there's data for this day
            # If there is, then go on to determine what file it is #
            try:
                try:  # Prioritize reading what's in the crawler
                    # Read Excel or prioritize location then extract info
                    each_path = f"{path_cta}/{users}/{difference}.xls"
                    table = pd.read_excel(each_path)
                    # First, determine the location.
                    where_is_settle = table["######"] == 'settle'
                    where_is_moneyin = table["#####"] == 'money in'
                    where_is_commission = table["#####"] == 'commission'
                    where_is_margin_occupied = table["#####"] == 'margin occupied'
                    pos_settle = np.flatnonzero(where_is_settle)[0]
                    pos_moneyin = np.flatnonzero(where_is_moneyin)[0]
                    pos_commission = np.flatnonzero(where_is_commission)[0]
                    pos_margin_occupied = np.flatnonzero(where_is_margin_occupied)[0]
                    settle += table["#####"].loc[pos_settle]
                    money_in += table["#####"].loc[pos_moneyin]
                    commission += table["#####"].loc[pos_commission]
                    margin_occupied += table["#####"].loc[pos_margin_occupied]

                    # 除此之外，每个single_account的变量也要弄，每个single_account的权益、手续费、margin如下
                    today_set, today_com, today_mar = table["#####"].loc[pos_settle], \
                                                      table["#####"].loc[pos_commission], \
                                                      table["#####"].loc[pos_margin_occupied]
                    # And access is to add up the previous day's base, how to get it with openpyxl is said here again:
                    # list(table.rows) is a big list containing information about each row, 
                    # and each element in that big list # is a tuple about each row, 
                    # and each element in that tuple is a tuple about each row.
                    # Each element in the list is a tuple about each row, and each element in the tuple is in the form of a cell. 
                    # cell.value is the original value.
                    # For example, if I want to get information about row i and column j, it is list(table.rows)[i][j].value
                    today_mon = float(list(write_main_table.rows)[last_date_pos][header.index(users) + 1].value) + \
                                table["#####"].loc[pos_moneyin]
                except:
                    # Don't read the account_summary if there is a problem, report an error.
                    # The error reporting mechanism is variable referenced before assignment.
                    # Because this exception does not define these variables, but they will be used later, 
                    # the variable undefined will be present as soon as the exception is entered.
                    print(f"{users} doesn't have today's date")
                    raise Exception
            # If it's not there, it's probably missing data and can't be passed directly to the next user.
            except:
                today_set, today_com, today_mar = 0, 0, 0
                # Also, if it's open but no action, then the access totals will be based on the previous day's
                if list(write_main_table.rows)[last_date_pos][header.index(users) + 1].value != None:
                    today_mon = float(list(write_main_table.rows)[last_date_pos][header.index(users) + 1].value)
                # if just not open, set to 0
                else:
                    today_mon = 0

            informations[header.index(users)] = float(format(today_set, ".2f"))
            informations[header.index(users) + 1] = float(format(today_mon, ".2f"))
            informations[header.index(users) + 2] = today_com
            informations[header.index(users) + 3] = today_mar
        date_to_write = "/".join([str(int(num)) for num in difference.split("-")])
        informations = [date_to_write, float(format(settle, ".2f")), float(format(money_in, ".2f")),
                        commission, margin_occupied / settle, ""] + informations[6:]
        for i in range(1, len(informations) + 1):
            # here last_pos to add 2, because first of all Excel rows and columns are from the beginning of 1 instead of 0, 
            # and secondly to write to the original file last line of the next line
            write_main_table.cell(last_date_pos + 2, i, informations[i - 1])
        # you must save it
        write_main_table_book.save(file_path)
        # and lastly update the postition of the last row
        last_date_pos += 1



c_path = "/".join(os.path.abspath(__file__).split("/")[:-1])
r_path = "/Volumes/#####/#####data"
s_path = "/Volumes/#####/#####data"
update("########", c_path, r_path, s_path)


